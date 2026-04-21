"""
=============================================================================
  Product Screenshot Processor  —  v3.0  (Production Release)
=============================================================================
  Extract product names & prices from screenshots → export to Excel.

  Setup:
      pip install opencv-python pytesseract pandas pillow openpyxl

  Tesseract OCR engine (install separately):
      Windows : https://github.com/UB-Mannheim/tesseract/wiki
      macOS   : brew install tesseract
      Ubuntu  : sudo apt install tesseract-ocr
=============================================================================
"""

# ── stdlib ────────────────────────────────────────────────────────────────────
import sys, io, os, re, threading, queue
from datetime import datetime

# Force UTF-8 on Windows consoles
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# ── third-party ───────────────────────────────────────────────────────────────
import cv2
import pytesseract
import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk, ImageDraw
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# =============================================================================
#  CONFIGURATION  ←  edit here before running
# =============================================================================

# Windows example: r"C:\Program Files\Tesseract-OCR\tesseract.exe"
# macOS / Linux  : leave as None (uses system PATH)
TESSERACT_CMD = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

OCR_LANG = "eng"   # use "eng+ara" for Arabic+English

PRICE_PATTERNS = [
    r"(?:OMR|omr)\s*\d{1,4}(?:[.,]\d{1,3})?",
    r"\d{1,4}(?:[.,]\d{1,3})?\s*(?:OMR|omr)",
    r"\d{1,4}(?:[.,]\d{1,2})?\s*(?:AED|SAR|KWD|BHD|QAR)",
    r"(?:AED|SAR|KWD|BHD|QAR)\s*\d{1,4}(?:[.,]\d{1,2})?",
    r"[\$\£\€\¥]\s*\d{1,6}(?:[.,]\d{1,2})?",
    r"\b\d{1,4}[.,]\d{2,3}\b",
]

PRODUCT_CODE_RE = re.compile(
    r"\b[A-Za-z]{1,5}[-./][A-Za-z0-9]{1,5}[-./][A-Za-z0-9]{3,12}\b"
    r"|\b[A-Za-z]{1,4}[-./][A-Za-z0-9]{4,15}\b"
    r"|\b[A-Za-z0-9]{2,5}[-]{1,2}[0-9]{5,12}\b"
)

# =============================================================================
#  COLOUR PALETTE
# =============================================================================

C = {
    "bg":           "#0D0F18",
    "panel":        "#13151F",
    "card":         "#1C1F2E",
    "card2":        "#21253A",
    "border":       "#2A2D42",
    "border_light": "#3A3F5C",
    "accent":       "#5B8DEF",
    "accent_hover": "#7BA7FF",
    "accent2":      "#8B5CF6",
    "green":        "#10B981",
    "green_hover":  "#34D399",
    "red":          "#EF4444",
    "yellow":       "#F59E0B",
    "text":         "#CBD5E1",
    "text_dim":     "#4B5563",
    "text_muted":   "#6B7280",
    "text_bright":  "#F1F5F9",
    "white":        "#FFFFFF",
}


# =============================================================================
#  CV / OCR  UTILITIES
# =============================================================================

def configure_tesseract() -> bool:
    if TESSERACT_CMD:
        pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD
    try:
        pytesseract.get_tesseract_version()
        return True
    except pytesseract.TesseractNotFoundError:
        return False


def load_image(path: str):
    img = cv2.imread(path)
    if img is None:
        raise FileNotFoundError(f"Cannot open: {path}")
    return img


def preprocess_for_ocr(crop_bgr) -> list:
    """Return multiple pre-processed PIL variants for best OCR coverage."""
    h, w = crop_bgr.shape[:2]
    scale = 1.0
    if max(h, w) < 600:
        scale = max(2.0, 600 / max(h, w))
    if scale > 1.0:
        crop_bgr = cv2.resize(crop_bgr,
                              (int(w * scale), int(h * scale)),
                              interpolation=cv2.INTER_CUBIC)

    gray = cv2.cvtColor(crop_bgr, cv2.COLOR_BGR2GRAY)
    variants = []

    # 1. OTSU
    _, otsu = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    variants.append(Image.fromarray(otsu))

    # 2. OTSU inverted  (light text on dark bg)
    variants.append(Image.fromarray(cv2.bitwise_not(otsu)))

    # 3. Adaptive threshold  (uneven lighting)
    ada = cv2.adaptiveThreshold(gray, 255,
                                cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                cv2.THRESH_BINARY, 31, 10)
    variants.append(Image.fromarray(ada))

    # 4. Sharpened gray
    denoised = cv2.fastNlMeansDenoising(gray, h=10)
    sharp = cv2.addWeighted(denoised, 1.8,
                            cv2.GaussianBlur(denoised, (3, 3), 0), -0.8, 0)
    variants.append(Image.fromarray(sharp))

    # 5. CLAHE + 3x upscale  (product codes with hyphens)
    clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
    eq = clahe.apply(gray)
    _, eq_bin = cv2.threshold(eq, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    big = cv2.resize(eq_bin,
                     (eq_bin.shape[1] * 3, eq_bin.shape[0] * 3),
                     interpolation=cv2.INTER_NEAREST)
    variants.append(Image.fromarray(big))

    return variants


def run_ocr(pil_images) -> str:
    """Run Tesseract on all variants x all PSM modes; return best result."""
    psm_modes = [
        "--oem 3 --psm 6",
        "--oem 3 --psm 11",
        "--oem 3 --psm 3",
        "--oem 3 --psm 7",
        "--oem 3 --psm 8",
    ]
    whitelist = (
        r"-c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        r"abcdefghijklmnopqrstuvwxyz0123456789-./() "
    )
    if not isinstance(pil_images, list):
        pil_images = [pil_images]

    best_text, best_score = "", -1
    for pil in pil_images:
        for psm in psm_modes:
            cfg = f"{psm} -l {OCR_LANG} {whitelist}"
            try:
                text = pytesseract.image_to_string(pil, config=cfg)
            except Exception:
                continue
            score = (sum(1 for c in text if c.isdigit()) * 3 +
                     sum(1 for c in text if c.isalnum()))
            if score > best_score:
                best_score, best_text = score, text

    return best_text


def clean_text(raw: str) -> str:
    text = re.sub(r"[ \t]+", " ", raw)
    lines = [ln.strip() for ln in text.splitlines()
             if len(ln.strip()) > 2 and
             not re.fullmatch(r"[^\w\s]+", ln.strip())]
    return "\n".join(lines).strip()


def extract_price(text: str) -> str:
    for p in PRICE_PATTERNS:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            return m.group(0).strip()
    return ""


def extract_product_name(text: str) -> str:
    # Priority 1 — product code  (e.g. A-M-2510520)
    m = PRODUCT_CODE_RE.search(text)
    if m:
        return m.group(0).strip().upper()

    # Priority 2 — longest clean line after removing prices
    clean = text
    for p in PRICE_PATTERNS:
        clean = re.sub(p, "", clean, flags=re.IGNORECASE)

    lines = [ln.strip() for ln in clean.splitlines() if len(ln.strip()) > 2]
    if not lines:
        return ""

    candidate = max(lines[:5], key=len)
    if candidate.islower():
        candidate = candidate.upper()
    candidate = re.sub(r"^[\s|:،,.\-]+", "", candidate)
    candidate = re.sub(r"[\s|:،,.\-]+$", "", candidate)
    return candidate.strip()


# =============================================================================
#  EXCEL EXPORT
# =============================================================================

def export_to_excel(results: list, output_path: str) -> str:
    df = pd.DataFrame(results)[["file", "product_name", "price"]]
    df.columns = ["Source File", "Product Name", "Price"]
    df.to_excel(output_path, index=False, sheet_name="Products")

    wb = load_workbook(output_path)
    ws = wb.active

    HDR_FILL = PatternFill("solid", fgColor="1E3A5F")
    ALT_FILL = PatternFill("solid", fgColor="F0F4FF")
    HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    DAT_FONT = Font(name="Calibri", size=10)
    C_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    L_ALIGN  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    THIN     = Side(style="thin", color="CCCCCC")
    BDR      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    for cell in ws[1]:
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = C_ALIGN; cell.border = BDR

    for ri, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = ALT_FILL if ri % 2 == 0 else PatternFill()
        for cell in row:
            cell.font = DAT_FONT; cell.fill = fill
            cell.alignment = L_ALIGN; cell.border = BDR

    for col, width in zip(["A", "B", "C"], [32, 50, 20]):
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 24
    for ri in range(2, ws.max_row + 1):
        ws.row_dimensions[ri].height = 20

    sr = ws.max_row + 2
    ws.cell(sr, 1, "Total products:").font = Font(name="Calibri", bold=True, size=10)
    ws.cell(sr, 2, f"=COUNTA(B2:B{ws.max_row - 1})").font = Font(name="Calibri", bold=True, size=10)

    wb.save(output_path)
    return output_path


# =============================================================================
#  MAIN APPLICATION
# =============================================================================

class App(tk.Tk):
    """Product Screenshot Processor — production GUI."""

    # =========================================================================
    #  INIT
    # =========================================================================

    def __init__(self):
        super().__init__()
        self.title("Product Screenshot Processor")
        self.geometry("1360x820")
        self.minsize(1100, 700)
        self.configure(bg=C["bg"])

        self._apply_styles()

        # application state
        self.image_paths: list = []
        self.current_idx: int  = -1
        self.current_img       = None
        self.current_box       = None
        self.results:     list = []

        # canvas rendering
        self.photo_ref         = None
        self.scale_x           = 1.0
        self.scale_y           = 1.0
        self.img_offset_x      = 0
        self.img_offset_y      = 0

        # draw-crop state
        self.draw_mode         = False
        self.rect_start        = None
        self.rect_id           = None

        # zoom / pan
        self.zoom_level        = 1.0
        self.zoom_min          = 0.15
        self.zoom_max          = 10.0
        self.pan_start         = None
        self.pan_offset_x      = 0
        self.pan_offset_y      = 0

        # async OCR
        self._ocr_q: queue.Queue = queue.Queue()

        self._build_ui()
        self._check_tesseract()
        self.after(120, self._poll_ocr)

    # =========================================================================
    #  STYLES
    # =========================================================================

    def _apply_styles(self):
        s = ttk.Style(self)
        s.theme_use("clam")
        s.configure(".", background=C["bg"], foreground=C["text"],
                    font=("Segoe UI", 10), borderwidth=0, relief="flat")
        s.configure("TFrame", background=C["bg"])
        s.configure("TLabel", background=C["bg"], foreground=C["text"])
        s.configure("TProgressbar",
                    troughcolor=C["border"], background=C["accent"],
                    lightcolor=C["accent"], darkcolor=C["accent"],
                    borderwidth=0, thickness=4)
        s.configure("Vertical.TScrollbar",
                    background=C["border"], troughcolor=C["panel"],
                    arrowcolor=C["text_dim"], borderwidth=0, width=6)
        s.map("Vertical.TScrollbar",
              background=[("active", C["accent"])])
        s.configure("Tree.Treeview",
                    background=C["card"], foreground=C["text"],
                    fieldbackground=C["card"], rowheight=32,
                    font=("Segoe UI", 9), borderwidth=0)
        s.configure("Tree.Treeview.Heading",
                    background=C["card2"], foreground=C["text_bright"],
                    font=("Segoe UI", 9, "bold"), relief="flat",
                    padding=(8, 6))
        s.map("Tree.Treeview",
              background=[("selected", C["accent"])],
              foreground=[("selected", C["white"])])

    # =========================================================================
    #  UI LAYOUT
    # =========================================================================

    def _build_ui(self):
        self._build_titlebar()
        body = tk.Frame(self, bg=C["bg"])
        body.pack(fill="both", expand=True)
        self._build_sidebar(body)
        self._build_canvas_area(body)
        self._build_results_panel(body)
        self._build_statusbar()

    # ── Title bar ─────────────────────────────────────────────────────────────

    def _build_titlebar(self):
        bar = tk.Frame(self, bg=C["panel"], height=52)
        bar.pack(fill="x")
        bar.pack_propagate(False)
        tk.Frame(bar, bg=C["border"], height=1).pack(fill="x", side="bottom")

        tk.Label(bar, text="  Product Processor",
                 bg=C["panel"], fg=C["text_bright"],
                 font=("Segoe UI", 13, "bold")).pack(side="left", padx=(16, 4))
        tk.Label(bar, text="v3.0",
                 bg=C["panel"], fg=C["text_dim"],
                 font=("Segoe UI", 9)).pack(side="left")

        self.tess_badge = tk.Label(
            bar, text=" Tesseract... ",
            bg=C["yellow"], fg="#111",
            font=("Segoe UI", 8, "bold"), padx=10, pady=4)
        self.tess_badge.pack(side="right", padx=12, pady=13)

        self._mk_btn(bar, "Export Excel", self._export,
                     C["green"], C["green_hover"],
                     fg="#0a0a0a").pack(side="right", pady=12, padx=(0, 4))

    # ── Sidebar ───────────────────────────────────────────────────────────────

    def _build_sidebar(self, parent):
        sb = tk.Frame(parent, bg=C["panel"], width=210)
        sb.pack(side="left", fill="y")
        sb.pack_propagate(False)
        tk.Frame(sb, bg=C["border"], width=1).pack(side="right", fill="y")

        tk.Label(sb, text="IMAGES", bg=C["panel"], fg=C["text_dim"],
                 font=("Segoe UI", 8, "bold")).pack(
                     anchor="w", padx=16, pady=(16, 6))

        self._mk_btn(sb, "+ Load Images", self._load_images,
                     C["accent"], C["accent_hover"]).pack(
                         fill="x", padx=12, pady=(0, 10), ipady=2)

        lf = tk.Frame(sb, bg=C["panel"])
        lf.pack(fill="both", expand=True, padx=12)
        vsb = ttk.Scrollbar(lf, orient="vertical",
                             style="Vertical.TScrollbar")
        vsb.pack(side="right", fill="y")
        self.img_listbox = tk.Listbox(
            lf, bg=C["card"], fg=C["text"],
            selectbackground=C["accent"], selectforeground=C["white"],
            relief="flat", bd=0, font=("Segoe UI", 9),
            activestyle="none", highlightthickness=0,
            yscrollcommand=vsb.set)
        self.img_listbox.pack(fill="both", expand=True)
        vsb.config(command=self.img_listbox.yview)
        self.img_listbox.bind("<<ListboxSelect>>", self._on_list_select)

        self.list_lbl = tk.Label(sb, text="No images loaded",
                                  bg=C["panel"], fg=C["text_dim"],
                                  font=("Segoe UI", 8), wraplength=185)
        self.list_lbl.pack(pady=(6, 16), padx=12)

    # ── Canvas area ───────────────────────────────────────────────────────────

    def _build_canvas_area(self, parent):
        centre = tk.Frame(parent, bg=C["bg"])
        centre.pack(side="left", fill="both", expand=True)

        # Canvas container
        cf = tk.Frame(centre, bg=C["card"],
                      highlightbackground=C["border"], highlightthickness=1)
        cf.pack(fill="both", expand=True, padx=(10, 6), pady=(10, 6))

        self.canvas = tk.Canvas(cf, bg=C["card"],
                                cursor="crosshair", highlightthickness=0)
        self.canvas.pack(fill="both", expand=True)

        # Mouse bindings
        self.canvas.bind("<ButtonPress-1>",   self._on_press)
        self.canvas.bind("<B1-Motion>",        self._on_drag)
        self.canvas.bind("<ButtonRelease-1>", self._on_release)
        self.canvas.bind("<MouseWheel>",       self._on_zoom)
        self.canvas.bind("<Button-4>",         self._on_zoom)
        self.canvas.bind("<Button-5>",         self._on_zoom)
        self.canvas.bind("<ButtonPress-2>",    self._on_pan_start)
        self.canvas.bind("<B2-Motion>",        self._on_pan_move)
        self.canvas.bind("<ButtonRelease-2>",  self._on_pan_end)

        # Toolbar row
        tb = tk.Frame(centre, bg=C["bg"])
        tb.pack(fill="x", padx=(10, 6), pady=(0, 6))

        self._mk_btn(tb, "Accept", self._accept,
                     C["green"], C["green_hover"],
                     fg="#0a0a0a", width=12).pack(side="left", padx=(0, 5))
        self._mk_btn(tb, "Redraw", self._start_draw,
                     C["yellow"], "#FBBF24",
                     fg="#0a0a0a", width=10).pack(side="left", padx=(0, 5))
        self._mk_btn(tb, "Skip", self._skip,
                     C["red"], "#F87171", width=8).pack(side="left")

        for txt, cmd in [("Next", self._next_image),
                         ("Prev", self._prev_image)]:
            self._mk_btn(tb, txt, cmd, C["card2"], C["border"],
                         width=6).pack(side="right", padx=(4, 0))

        tk.Frame(tb, bg=C["border"], width=1).pack(
            side="right", fill="y", padx=8, pady=4)

        for txt, cmd in [("Reset", self._zoom_reset),
                         (" - ",   self._zoom_out),
                         (" + ",   self._zoom_in)]:
            self._mk_btn(tb, txt, cmd, C["card2"], C["border"],
                         width=5).pack(side="right", padx=(4, 0))

        # OCR fields
        of = tk.Frame(centre, bg=C["panel"],
                      highlightbackground=C["border"], highlightthickness=1)
        of.pack(fill="x", padx=(10, 6), pady=(0, 10))

        ocr_top = tk.Frame(of, bg=C["panel"])
        ocr_top.pack(fill="x", padx=12, pady=(8, 4))
        tk.Label(ocr_top, text="EXTRACTED DATA",
                 bg=C["panel"], fg=C["text_dim"],
                 font=("Segoe UI", 8, "bold")).pack(side="left")
        self.ocr_lbl = tk.Label(ocr_top, text="",
                                 bg=C["panel"], fg=C["accent"],
                                 font=("Segoe UI", 8, "italic"))
        self.ocr_lbl.pack(side="left", padx=(8, 0))

        fields = tk.Frame(of, bg=C["panel"])
        fields.pack(fill="x", padx=12, pady=(0, 8))
        fields.columnconfigure(1, weight=1)

        self.name_var  = tk.StringVar()
        self.price_var = tk.StringVar()

        for row, (lbl, var) in enumerate([
                ("Product Name", self.name_var),
                ("Price",        self.price_var)]):
            tk.Label(fields, text=lbl, bg=C["panel"], fg=C["text_muted"],
                     font=("Segoe UI", 9), width=14,
                     anchor="w").grid(row=row, column=0, sticky="w", pady=4)
            tk.Entry(fields, textvariable=var,
                     bg=C["card2"], fg=C["text_bright"],
                     insertbackground=C["text"], relief="flat",
                     font=("Segoe UI", 10),
                     highlightbackground=C["border_light"],
                     highlightthickness=1).grid(
                         row=row, column=1, sticky="ew",
                         padx=(8, 0), pady=4, ipady=6)

        self._mk_btn(of, "+ Add to Results", self._add_result,
                     C["accent2"], "#9D6EFF",
                     width=18).pack(anchor="e", padx=12, pady=(0, 10))

    # ── Results panel ─────────────────────────────────────────────────────────

    def _build_results_panel(self, parent):
        rp = tk.Frame(parent, bg=C["panel"], width=310)
        rp.pack(side="right", fill="y")
        rp.pack_propagate(False)
        tk.Frame(rp, bg=C["border"], width=1).pack(side="left", fill="y")

        inner = tk.Frame(rp, bg=C["panel"])
        inner.pack(fill="both", expand=True, padx=(12, 10))

        # Stat cards
        stats = tk.Frame(inner, bg=C["panel"])
        stats.pack(fill="x", pady=(14, 10))
        self._stat_proc = self._mk_stat(stats, "0", "Processed")
        self._stat_proc.pack(side="left", fill="x", expand=True, padx=(0, 5))
        self._stat_extr = self._mk_stat(stats, "0", "Extracted")
        self._stat_extr.pack(side="left", fill="x", expand=True)

        tk.Label(inner, text="RESULTS", bg=C["panel"], fg=C["text_dim"],
                 font=("Segoe UI", 8, "bold")).pack(anchor="w", pady=(0, 4))

        # Treeview
        tf = tk.Frame(inner, bg=C["panel"])
        tf.pack(fill="both", expand=True)
        tv_vsb = ttk.Scrollbar(tf, orient="vertical",
                                style="Vertical.TScrollbar")
        tv_vsb.pack(side="right", fill="y")
        self.tree = ttk.Treeview(tf, columns=("name", "price"),
                                  show="headings", style="Tree.Treeview",
                                  yscrollcommand=tv_vsb.set)
        self.tree.heading("name",  text="Product Name")
        self.tree.heading("price", text="Price")
        self.tree.column("name",  width=190, anchor="w")
        self.tree.column("price", width=80,  anchor="center")
        self.tree.pack(fill="both", expand=True)
        tv_vsb.config(command=self.tree.yview)

        # Right-click context menu
        ctx = tk.Menu(self, tearoff=0, bg=C["card"], fg=C["text"],
                      activebackground=C["accent"],
                      activeforeground=C["white"],
                      font=("Segoe UI", 9), bd=0, relief="flat")
        ctx.add_command(label="Edit row",   command=self._edit_row)
        ctx.add_separator()
        ctx.add_command(label="Delete row", command=self._delete_row)
        self.tree.bind("<Button-3>",
                       lambda e: ctx.tk_popup(e.x_root, e.y_root))

        tk.Frame(inner, bg=C["border"], height=1).pack(
            fill="x", pady=(10, 8))
        self._mk_btn(inner, "Export to Excel", self._export,
                     C["green"], C["green_hover"],
                     fg="#0a0a0a", height=2).pack(fill="x", pady=(0, 5))
        self._mk_btn(inner, "Clear All", self._clear_all,
                     C["card2"], C["border"]).pack(fill="x")

    # ── Status bar ────────────────────────────────────────────────────────────

    def _build_statusbar(self):
        sb = tk.Frame(self, bg=C["panel"], height=28)
        sb.pack(fill="x", side="bottom")
        sb.pack_propagate(False)
        tk.Frame(sb, bg=C["border"], height=1).pack(fill="x", side="top")
        self.status_var = tk.StringVar(value="Ready — load images to begin.")
        tk.Label(sb, textvariable=self.status_var,
                 bg=C["panel"], fg=C["text_muted"],
                 font=("Segoe UI", 8), anchor="w").pack(
                     side="left", padx=14, pady=4)
        self.progress = ttk.Progressbar(sb, mode="determinate",
                                         style="TProgressbar", length=160)
        self.progress.pack(side="right", padx=14, pady=8)

    # =========================================================================
    #  WIDGET FACTORIES
    # =========================================================================

    def _mk_btn(self, parent, text, cmd, bg, hover,
                fg=None, width=None, height=1):
        fg = fg or C["text_bright"]
        kw = dict(text=text, command=cmd, bg=bg, fg=fg,
                  relief="flat", font=("Segoe UI", 9, "bold"),
                  cursor="hand2", activebackground=hover,
                  activeforeground=fg, pady=6, padx=12,
                  height=height, bd=0)
        if width:
            kw["width"] = width
        b = tk.Button(parent, **kw)
        b.bind("<Enter>", lambda e: b.config(bg=hover))
        b.bind("<Leave>", lambda e: b.config(bg=bg))
        return b

    def _mk_stat(self, parent, val, label):
        f = tk.Frame(parent, bg=C["card"],
                     highlightbackground=C["border"],
                     highlightthickness=1)
        n = tk.Label(f, text=val, bg=C["card"], fg=C["accent"],
                     font=("Segoe UI", 24, "bold"))
        n.pack(pady=(6, 0))
        tk.Label(f, text=label, bg=C["card"], fg=C["text_dim"],
                 font=("Segoe UI", 8)).pack(pady=(0, 6))
        f._num = n
        return f

    # =========================================================================
    #  TESSERACT CHECK
    # =========================================================================

    def _check_tesseract(self):
        ok = configure_tesseract()
        if ok:
            self.tess_badge.config(text=" Tesseract OK ",
                                   bg=C["green"], fg="#0a0a0a")
        else:
            self.tess_badge.config(text=" Tesseract Missing ",
                                   bg=C["red"], fg=C["white"])
            messagebox.showerror(
                "Tesseract Not Found",
                "Tesseract OCR engine not found.\n\n"
                "Windows: https://github.com/UB-Mannheim/tesseract/wiki\n"
                "         Set TESSERACT_CMD at the top of this script.\n\n"
                "macOS  : brew install tesseract\n"
                "Ubuntu : sudo apt install tesseract-ocr")

    # =========================================================================
    #  IMAGE LOADING & NAVIGATION
    # =========================================================================

    def _load_images(self):
        paths = filedialog.askopenfilenames(
            title="Select product screenshots",
            filetypes=[("Images",
                        "*.png *.jpg *.jpeg *.bmp *.tiff *.webp"),
                       ("All files", "*.*")])
        if not paths:
            return
        self.image_paths = list(paths)
        self.img_listbox.delete(0, "end")
        for p in self.image_paths:
            self.img_listbox.insert("end", "  " + os.path.basename(p))
        self.list_lbl.config(
            text=f"{len(self.image_paths)} image(s) loaded")
        self.progress["maximum"] = len(self.image_paths)
        self._set_status(f"Loaded {len(self.image_paths)} image(s).")
        self._goto(0)

    def _on_list_select(self, _event):
        sel = self.img_listbox.curselection()
        if sel:
            self._goto(sel[0])

    def _goto(self, idx: int):
        if not self.image_paths or not (0 <= idx < len(self.image_paths)):
            return
        self.current_idx = idx
        self.draw_mode = False
        self.img_listbox.selection_clear(0, "end")
        self.img_listbox.selection_set(idx)
        self.img_listbox.see(idx)
        self._open_image(self.image_paths[idx])

    def _prev_image(self):
        self._goto(self.current_idx - 1)

    def _next_image(self):
        self._goto(self.current_idx + 1)

    # =========================================================================
    #  CANVAS RENDERING
    # =========================================================================

    def _open_image(self, path: str):
        """Load image, reset view, start draw mode."""
        self.zoom_level   = 1.0
        self.pan_offset_x = 0
        self.pan_offset_y = 0
        self.current_box  = None
        self.name_var.set("")
        self.price_var.set("")
        self.ocr_lbl.config(text="")

        try:
            self.current_img = load_image(path)
        except FileNotFoundError as e:
            messagebox.showerror("Error", str(e))
            return

        self._redraw_canvas()
        self._set_status(
            f"Loaded: {os.path.basename(path)}  —  draw a crop to begin.")
        self._start_draw()

    def _redraw_canvas(self):
        """Render current image at current zoom/pan with optional crop box."""
        if self.current_img is None:
            return

        self.canvas.update_idletasks()
        cw = self.canvas.winfo_width()  or 700
        ch = self.canvas.winfo_height() or 500

        h_img, w_img = self.current_img.shape[:2]
        base  = min(cw / w_img, ch / h_img, 1.0)
        scale = base * self.zoom_level
        nw    = max(1, int(w_img * scale))
        nh    = max(1, int(h_img * scale))

        self.scale_x      = scale
        self.scale_y      = scale
        self.img_offset_x = (cw - nw) // 2 + self.pan_offset_x
        self.img_offset_y = (ch - nh) // 2 + self.pan_offset_y

        rgb = cv2.cvtColor(self.current_img, cv2.COLOR_BGR2RGB)
        pil = Image.fromarray(rgb).resize((nw, nh), Image.LANCZOS)

        # Draw bounding box
        if self.current_box:
            draw = ImageDraw.Draw(pil)
            x, y, w, h = self.current_box
            px = int(x * scale); py = int(y * scale)
            pw = int(w * scale); ph = int(h * scale)
            for t, col in [(3, "#1E3A5F"), (2, "#3B7DE8"), (1, "#7BA7FF")]:
                draw.rectangle(
                    [px - t, py - t, px + pw + t, py + ph + t],
                    outline=col)
            bw = 52
            draw.rectangle([px, py - 20, px + bw, py], fill="#5B8DEF")
            draw.text((px + 4, py - 17), " CROP ", fill=C["white"])

        photo = ImageTk.PhotoImage(pil)
        self.photo_ref = photo

        self.canvas.delete("all")
        self.canvas.create_image(self.img_offset_x, self.img_offset_y,
                                  anchor="nw", image=photo)
        # Zoom % badge
        self.canvas.create_text(
            cw - 10, ch - 10,
            text=f"{int(self.zoom_level * 100)}%",
            anchor="se", fill=C["text_dim"],
            font=("Segoe UI", 9, "bold"))

    # =========================================================================
    #  DRAW CROP
    # =========================================================================

    def _start_draw(self):
        self.draw_mode  = True
        self.rect_start = None
        self.rect_id    = None
        self.canvas.config(cursor="crosshair")
        self._set_status("Draw a crop rectangle around the product.")

    def _on_press(self, event):
        if not self.draw_mode:
            return
        self.rect_start = (event.x, event.y)
        if self.rect_id:
            self.canvas.delete(self.rect_id)

    def _on_drag(self, event):
        if not self.draw_mode or not self.rect_start:
            return
        if self.rect_id:
            self.canvas.delete(self.rect_id)
        x0, y0 = self.rect_start
        self.rect_id = self.canvas.create_rectangle(
            x0, y0, event.x, event.y,
            outline=C["yellow"], width=2, dash=(5, 3))

    def _on_release(self, event):
        if not self.draw_mode or not self.rect_start:
            return
        self.draw_mode = False
        self.canvas.config(cursor="crosshair")
        if self.rect_id:
            self.canvas.delete(self.rect_id)
            self.rect_id = None

        x0, y0 = self.rect_start
        x1, y1 = event.x, event.y

        cx0 = (min(x0, x1) - self.img_offset_x) / self.scale_x
        cy0 = (min(y0, y1) - self.img_offset_y) / self.scale_y
        cx1 = (max(x0, x1) - self.img_offset_x) / self.scale_x
        cy1 = (max(y0, y1) - self.img_offset_y) / self.scale_y

        h_img, w_img = self.current_img.shape[:2]
        cx0 = max(0, int(cx0)); cy0 = max(0, int(cy0))
        cx1 = min(w_img, int(cx1)); cy1 = min(h_img, int(cy1))

        if cx1 - cx0 < 10 or cy1 - cy0 < 10:
            self._set_status("Selection too small — draw again.")
            self._start_draw()
            return

        self.current_box = (cx0, cy0, cx1 - cx0, cy1 - cy0)
        self._redraw_canvas()
        self._launch_ocr()

    # =========================================================================
    #  ZOOM & PAN
    # =========================================================================

    def _on_zoom(self, event):
        if self.current_img is None:
            return

        factor = 0.85 if (event.num == 5 or event.delta < 0) else 1.18
        new_z  = max(self.zoom_min, min(self.zoom_max, self.zoom_level * factor))

        # حساب الأبعاد الجديدة بعد الزوم
        cw = self.canvas.winfo_width()  or 700
        ch = self.canvas.winfo_height() or 500
        h_img, w_img = self.current_img.shape[:2]
        base      = min(cw / w_img, ch / h_img, 1.0)
        new_scale = base * new_z
        new_nw    = max(1, int(w_img * new_scale))
        new_nh    = max(1, int(h_img * new_scale))

        # موضع المؤشر في إحداثيات الصورة الأصلية
        img_x = (event.x - self.img_offset_x) / self.scale_x
        img_y = (event.y - self.img_offset_y) / self.scale_y

        # pan جديد بحيث تبقى نقطة الصورة تحت المؤشر ثابتة
        self.pan_offset_x = int(event.x - img_x * new_scale - (cw - new_nw) // 2)
        self.pan_offset_y = int(event.y - img_y * new_scale - (ch - new_nh) // 2)

        self.zoom_level = new_z
        self._redraw_canvas()

    def _zoom_in(self):
        self.zoom_level = min(self.zoom_max, self.zoom_level * 1.3)
        self._redraw_canvas()

    def _zoom_out(self):
        self.zoom_level = max(self.zoom_min, self.zoom_level / 1.3)
        self._redraw_canvas()

    def _zoom_reset(self):
        self.zoom_level   = 1.0
        self.pan_offset_x = 0
        self.pan_offset_y = 0
        self._redraw_canvas()

    def _on_pan_start(self, event):
        self.pan_start = (event.x, event.y)

    def _on_pan_move(self, event):
        if not self.pan_start:
            return
        self.pan_offset_x += event.x - self.pan_start[0]
        self.pan_offset_y += event.y - self.pan_start[1]
        self.pan_start = (event.x, event.y)
        self._redraw_canvas()

    def _on_pan_end(self, _event):
        self.pan_start = None

    # =========================================================================
    #  ASYNC OCR
    # =========================================================================

    def _launch_ocr(self):
        if self.current_img is None or self.current_box is None:
            return
        self.ocr_lbl.config(text="Running OCR...")
        x, y, w, h = self.current_box
        crop = self.current_img[y: y + h, x: x + w].copy()

        def _worker():
            try:
                variants = preprocess_for_ocr(crop)
                raw      = run_ocr(variants)
                cleaned  = clean_text(raw)
                self._ocr_q.put(("ok",
                                  extract_product_name(cleaned),
                                  extract_price(cleaned)))
            except Exception as exc:
                self._ocr_q.put(("err", str(exc)))

        threading.Thread(target=_worker, daemon=True).start()

    def _poll_ocr(self):
        try:
            while True:
                item = self._ocr_q.get_nowait()
                if item[0] == "ok":
                    _, name, price = item
                    self.name_var.set(name)
                    self.price_var.set(price)
                    self.ocr_lbl.config(text="Done")
                    self._set_status(
                        f"OCR done  |  Name: {name[:40] or '(empty)'}  "
                        f"Price: {price or '—'}")
                else:
                    self.ocr_lbl.config(text="OCR error")
                    self._set_status(f"OCR error: {item[1]}")
        except queue.Empty:
            pass
        self.after(120, self._poll_ocr)

    # =========================================================================
    #  ACTIONS
    # =========================================================================

    def _accept(self):
        if self.current_img is not None:
            self._add_result()

    def _add_result(self):
        name  = self.name_var.get().strip()
        price = self.price_var.get().strip()
        if not name:
            messagebox.showwarning(
                "Empty Name",
                "Product name is empty.\n"
                "Edit the field or skip this image.")
            return
        fname = os.path.basename(self.image_paths[self.current_idx])
        self.results.append({"file": fname,
                              "product_name": name,
                              "price": price})
        self.tree.insert("", "end", values=(name, price or "—"))
        self._update_stats()
        self._set_status(f"Added: {name[:60]}")
        self._next_image()

    def _skip(self):
        if self.current_idx < 0:
            return
        self._set_status(
            f"Skipped: "
            f"{os.path.basename(self.image_paths[self.current_idx])}")
        self._next_image()

    def _edit_row(self):
        sel = self.tree.selection()
        if not sel:
            return
        idx   = self.tree.index(sel[0])
        nv    = tk.StringVar(value=self.results[idx]["product_name"])
        pv    = tk.StringVar(value=self.results[idx]["price"])

        win = tk.Toplevel(self)
        win.title("Edit Row")
        win.configure(bg=C["bg"])
        win.geometry("440x190")
        win.resizable(False, False)
        win.grab_set()

        for row, (lbl, var) in enumerate([("Product Name", nv),
                                           ("Price",        pv)]):
            tk.Label(win, text=lbl, bg=C["bg"], fg=C["text_muted"],
                     font=("Segoe UI", 9)).grid(
                         row=row, column=0, sticky="w",
                         padx=20, pady=(20 if row == 0 else 4, 4))
            tk.Entry(win, textvariable=var, bg=C["card2"],
                     fg=C["text_bright"], relief="flat",
                     font=("Segoe UI", 10),
                     highlightbackground=C["border_light"],
                     highlightthickness=1,
                     insertbackground=C["text"]).grid(
                         row=row, column=1, sticky="ew",
                         padx=(8, 20),
                         pady=(20 if row == 0 else 4, 4), ipady=6)
        win.columnconfigure(1, weight=1)

        def _save():
            self.results[idx]["product_name"] = nv.get().strip()
            self.results[idx]["price"]         = pv.get().strip()
            self.tree.item(sel[0],
                           values=(nv.get().strip(),
                                   pv.get().strip() or "—"))
            win.destroy()

        self._mk_btn(win, "Save", _save, C["accent"], C["accent_hover"],
                     width=12).grid(row=2, column=1, sticky="e",
                                    padx=(8, 20), pady=(12, 0))

    def _delete_row(self):
        sel = self.tree.selection()
        if not sel:
            return
        idx = self.tree.index(sel[0])
        self.tree.delete(sel[0])
        if 0 <= idx < len(self.results):
            self.results.pop(idx)
        self._update_stats()

    def _clear_all(self):
        if not self.results:
            return
        if messagebox.askyesno("Clear All",
                               "Delete all extracted results?"):
            self.results.clear()
            for item in self.tree.get_children():
                self.tree.delete(item)
            self._update_stats()

    # =========================================================================
    #  EXPORT
    # =========================================================================

    def _export(self):
        if not self.results:
            messagebox.showwarning("No Data", "No results to export yet.")
            return
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = filedialog.asksaveasfilename(
            title="Save Excel file",
            defaultextension=".xlsx",
            initialfile=f"products_{ts}.xlsx",
            filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        try:
            export_to_excel(self.results, path)
            self._set_status(
                f"Exported {len(self.results)} product(s) to "
                f"{os.path.basename(path)}")
            if messagebox.askyesno(
                    "Export Complete",
                    f"{len(self.results)} products saved.\n\nOpen file now?"):
                try:
                    if sys.platform == "win32":
                        os.startfile(path)
                    elif sys.platform == "darwin":
                        os.system(f'open "{path}"')
                    else:
                        os.system(f'xdg-open "{path}"')
                except Exception:
                    pass
        except Exception as exc:
            messagebox.showerror("Export Error", str(exc))

    # =========================================================================
    #  HELPERS
    # =========================================================================

    def _set_status(self, msg: str):
        self.status_var.set(msg)
        self.update_idletasks()

    def _update_stats(self):
        processed = self.current_idx + 1 if self.current_idx >= 0 else 0
        self._stat_proc._num.config(text=str(processed))
        self._stat_extr._num.config(text=str(len(self.results)))
        self.progress["value"] = processed


# =============================================================================
#  ENTRY POINT
# =============================================================================

if __name__ == "__main__":
    app = App()
    app.mainloop()
